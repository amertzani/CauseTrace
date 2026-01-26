"""
XLSX Agent
==========

Specialized agent for processing Excel XLSX files.
Uses openpyxl for Excel file processing.
"""

import os
from typing import Dict
from .base_agent import BaseAgent

# Try to import openpyxl
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XLSXAgent(BaseAgent):
    """Agent specialized in processing Excel XLSX files."""
    
    def __init__(self):
        super().__init__("XLSX Agent")
        self.supported_extensions = ['.xlsx', '.XLSX', '.xlsm', '.XLSM']
        if not OPENPYXL_AVAILABLE:
            print("⚠️  openpyxl not available. XLSX support disabled.")
    
    def can_process(self, file_path: str) -> bool:
        """Check if file is an XLSX and openpyxl is available."""
        if not OPENPYXL_AVAILABLE:
            return False
        ext = os.path.splitext(file_path)[1].lower()
        return ext in ['.xlsx', '.xlsm']
    
    def extract_text(self, file_path: str) -> str:
        """
        Extract text from XLSX file.
        Processes all sheets and formats data for knowledge extraction.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Formatted text representation of Excel data
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl library is not installed. Install it with: pip install openpyxl")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n=== Sheet: {sheet_name} ===")
                
                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row == 0 or max_col == 0:
                    text_parts.append("(Empty sheet)")
                    continue
                
                text_parts.append(f"Dimensions: {max_row} rows × {max_col} columns")
                
                # Extract header row (first row)
                headers = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"Column{col}")
                
                text_parts.append(f"Headers: {', '.join(headers)}")
                
                # Extract data rows (limit to first 50 rows per sheet for performance)
                text_parts.append("\n--- Sample Data ---")
                sample_size = min(50, max_row - 1)  # Exclude header row
                
                for row_idx in range(2, min(2 + sample_size, max_row + 1)):
                    row_data = []
                    for col_idx in range(1, max_col + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            row_data.append(f"{headers[col_idx - 1]}: {str(cell_value)}")
                    
                    if row_data:
                        text_parts.append(f"Row {row_idx - 1}: {' | '.join(row_data)}")
                
                if max_row - 1 > sample_size:
                    text_parts.append(f"... ({max_row - 1 - sample_size} more rows)")
            
            workbook.close()
            return "\n".join(text_parts).strip()
            
        except Exception as e:
            raise Exception(f"Error reading XLSX: {str(e)}")
    
    def extract_metadata(self, file_path: str) -> Dict:
        """Extract XLSX-specific metadata."""
        metadata = super().extract_metadata(file_path)
        
        if not OPENPYXL_AVAILABLE:
            return metadata
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            metadata['sheets'] = workbook.sheetnames
            metadata['sheet_count'] = len(workbook.sheetnames)
            
            # Get dimensions for each sheet
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info[sheet_name] = {
                    'rows': sheet.max_row,
                    'columns': sheet.max_column
                }
            metadata['sheet_info'] = sheet_info
            
            workbook.close()
        except Exception:
            metadata['sheets'] = []
            metadata['sheet_count'] = 0
        
        return metadata

